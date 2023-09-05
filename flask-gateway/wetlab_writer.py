import yaml
import json
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection
import os


def find_keys(node, kv):
    """
    generator to return all values of given key in dictionary
    :param node: a dictionary to be searched
    :param kv: the key to search for
    """
    if isinstance(node, list):
        for i in node:
            for x in find_keys(i, kv):
                yield x
    elif isinstance(node, dict):
        if kv in node:
            yield node[kv]
        for j in node.values():
            for x in find_keys(j, kv):
                yield x


def format_generator(object):
    return list(object)


def load_metadata(path):
    with open(path, "r") as f:
        data_dict = yaml.safe_load(f.read())
    return data_dict


def format_sample_names_search(search_result_samples, search_result_filenames):
    sample_id_ls = []
    sample_fullnames_ls = []
    sample_filenames_ls = []

    sample_dict = {}
    for entry in search_result_samples:
        if type(entry) == str:
            sample_id_ls.append(entry)
        else:
            sample_fullnames_ls.append(entry)
    sample_fullnames_ls = sum(sample_fullnames_ls, [])
    sample_filenames_ls = sum(search_result_filenames, [])
    for sample in sample_id_ls:
        sample_dict[sample] = []
        for sample_fullname, sample_filename in zip(sample_fullnames_ls, sample_filenames_ls):
            
            if sample in sample_fullname:
                sample_dict[sample].append([sample_fullname, sample_filename])
    return sample_dict


def format_conditions(all_conditions, keys):
    condition_dict = {}
    # neue ebene für organism key wegen exp setting
    for condition in all_conditions:
        condition_dict[condition["condition_name"]] = {}
        for key in keys:
            if key == "sample_name":
                condition_dict[condition["condition_name"]]["samples"] = format_sample_names_search(
                    format_generator(find_keys(condition, key)), format_generator(find_keys(condition, "filenames")))
            else:
                condition_dict[condition["condition_name"]][key] = format_generator(
                    find_keys(condition, key))
    return condition_dict

def write_wetlab(exp_setting_dict, filepath, filename):
    path = filepath + filename + "_wetlab.xlsx"
    excel_lock_pw = "sequencing"
    with open("config/wetlab_config.yaml", "r", encoding="utf-8") as f:
        wetlab_config = yaml.safe_load(f)
    wb = openpyxl.Workbook()
    # Fill first block blue
        # Define the blue color E6F1FF
    blue_fill = PatternFill(start_color='B6B6B6',
                        end_color='B6B6B6',
                        fill_type='solid')
    user_fill = PatternFill(start_color='F4A460',
                        end_color='F4A460',
                        fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    for exp_num, exp_setting in enumerate(exp_setting_dict, start=1):
        
        ws = wb.create_sheet(title=f'exp_{exp_num}')
        
        headers = ['sample_ID','condition','organism_name']
        additional_headers = []
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col).value = header
            ws.cell(row=2, column=col).border = Border(bottom=Side(style='medium'))
            ws.cell(row=2, column=col).font = bold_font
            ws.cell(row=2, column=col).alignment = align_center

        row_num = 3
        col_num = 4
        additional_headers_map = {}
        for sample_id, sample_data in exp_setting.items():
            for factor, factor_data in sample_data["factors"].items():
                if factor not in additional_headers:
                    additional_headers.append(factor)
                    additional_headers_map[factor] = col_num
                    ws.cell(row=2, column=col_num).value ="factor." + factor
                    ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
                    ws.cell(row=2, column=col_num).font = bold_font
                    ws.cell(row=2, column=col_num).alignment = align_center
                    col_num += 1
        
        additional_headers_map["sample_filename"] = col_num
        ws.cell(row=2, column=col_num).value ="sample_filename"
        ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
        ws.cell(row=2, column=col_num).font = bold_font
        ws.cell(row=2, column=col_num).alignment = align_center
        col_num += 1            
        additional_headers_map["sample_label"] = col_num
        ws.cell(row=2, column=col_num).value ="sample_label"
        ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
        ws.cell(row=2, column=col_num).font = bold_font
        ws.cell(row=2, column=col_num).alignment = align_center
        ws.cell(row=2, column=col_num).fill = user_fill
        col_num += 1
        additional_headers_map["condition_label"] = col_num
        ws.cell(row=2, column=col_num).value ="condition_label"
        ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
        ws.cell(row=2, column=col_num).font = bold_font
        ws.cell(row=2, column=col_num).alignment = align_center
        col_num += 1
        additional_headers_map["tube_label"] = col_num
        ws.cell(row=2, column=col_num).value ="tube_label"
        ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
        ws.cell(row=2, column=col_num).font = bold_font
        ws.cell(row=2, column=col_num).alignment = align_center
        col_num += 1       

        # Fill the table with dta for each sample
        for sample_id, sample_data in exp_setting.items():
            ws.cell(row=row_num, column=1).value = sample_id
            ws.cell(row=row_num, column=2).value = sample_data["condition"]
            ws.cell(row=row_num, column=3).value = sample_data["organism_name"]
            ws.cell(row=row_num, column=additional_headers_map["sample_filename"]).value = sample_data["sample_filename"]            
            ws.cell(row=row_num, column=additional_headers_map["sample_label"]).value = sample_data["sample_label"]
            ws.cell(row=row_num, column=additional_headers_map["condition_label"]).value = sample_data["condition_label"]
            ws.cell(row=row_num, column=additional_headers_map["tube_label"]).value = row_num-2

            for factor, factor_data in sample_data["factors"].items():
                ws.cell(row=row_num, column=additional_headers_map[factor]).value = str(factor_data)

            row_num += 1

        
        # Create a new alignment object with text wrapping 
        #wrap_alignment = Alignment(wrap_text=True)
        
        #Fill User input block
        align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in range(1, row_num):
            if row <= 2:
                for col in range(1, col_num):
                    ws.cell(row=row, column=col).fill = blue_fill

        user_columns = col_num
        
        #no idea where to put that...
        for crow in range(2): 
            ws.cell(row=crow+1, column=user_columns-2).fill = user_fill
            ws.cell(row=crow+1, column=user_columns-3).fill = user_fill
        ws.cell(row=2, column=user_columns-1).alignment = align_center_wrap

        for category_title, category_data in wetlab_config.items():
            color_fill = PatternFill(start_color=category_data["color"],
                            end_color=category_data["color"],
                            fill_type='solid')
            ws.cell(row=1, column=col_num).value = category_title
            ws.cell(row=1, column=col_num).border = Border(left=Side(style='medium'))
            ws.cell(row=1, column=col_num).font = bold_font
            ws.cell(row=1, column=col_num).alignment = align_center_wrap
            
            for row in range(3, row_num):
                ws.cell(row=row, column=col_num).border = Border(left=Side(style='medium'))
                
            #print(category_title, category_data)
            for i,header_object in enumerate(category_data["header_objects"]):
                print(i, header_object)
                ws.cell(row=2, column=col_num).value = header_object["header"]
                ws.cell(row=2, column=col_num).font = bold_font
                ws.cell(row=2, column=col_num).alignment = align_center_wrap
                ws.cell(row=2, column=col_num).fill = color_fill
                ws.cell(row=1, column=col_num).fill = color_fill
                #ws.cell(row=2, column=col_num).alignment = wrap_alignment
                if i==0:
                    ws.cell(row=2, column=col_num).border = Border(left=Side(style='medium'), bottom=Side(style='medium'))
                else:
                    ws.cell(row=2, column=col_num).border = Border(bottom=Side(style='medium'))
                
                #add validation for each cell
                if header_object["validation_type"] == "dropdown_stop":
                    valid_values = header_object["validation_values"]
                    validation_dropdown = DataValidation(type="list", formula1='"{}"'.format(','.join(valid_values)), showDropDown=False, errorStyle="stop", errorTitle="Invalid value entered", showErrorMessage=True)
                    ws.add_data_validation(validation_dropdown)
                    for row in range(3, row_num):
                        validation_dropdown.add(ws.cell(row=row, column=col_num))
                if header_object["validation_type"] == "dropdown":
                    valid_values = header_object["validation_values"]
                    validation_dropdown = DataValidation(type="list", formula1='"{}"'.format(','.join(valid_values)), showDropDown=False, errorStyle="warning", errorTitle="Invalid value entered", showErrorMessage=True, error = 'Not from list. Continue?')
                    ws.add_data_validation(validation_dropdown)
                    for row in range(3, row_num):
                        validation_dropdown.add(ws.cell(row=row, column=col_num))
                if header_object["validation_type"] == "between":
                    min_value, max_value = header_object["validation_values"]
                    validation_between = DataValidation(type="decimal", operator="between",formula1=min_value, formula2=max_value, errorStyle="stop", errorTitle="Invalid value entered", showErrorMessage=True)
                    ws.add_data_validation(validation_between)
                    for row in range(3, row_num):
                        validation_between.add(ws.cell(row=row, column=col_num))
                if header_object["validation_type"] == "formula":
                    min_value, max_value = header_object["validation_values"]
                    validation_between = DataValidation(type="decimal", operator="between",formula1=min_value, formula2=max_value, errorStyle="stop", errorTitle="Invalid value entered", showErrorMessage=True)
                    ws.add_data_validation(validation_between)
                    for row in range(3, row_num):
                        validation_between.add(ws.cell(row=row, column=col_num))
                
                #unlock user cells
                for row in range(3, row_num):
                    ws.cell(row=row, column=col_num).protection = Protection(locked=False)
                col_num += 1

        for col_index,col in enumerate(ws.columns):
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length and col_index+1 < user_columns:
                        #fixed size for all non-user cells
                        max_length = 10 #len(cell.value)
                        #small size for all numbers
                        if col_index+1 == user_columns-1:
                            max_length = 4
                        #bigger size for condition and sample name / remove protection
                        if col_index+1 == user_columns-2 or col_index+1 == user_columns-3:
                            max_length = 18
                            if col_index+1 == user_columns-3:
                                man_formula="=AND("
                                #only a-z A-Z 0-9 - _
                                man_formula+="ISNUMBER(SUMPRODUCT(SEARCH(MID("+col[col_index+1].column_letter+"3,ROW(INDIRECT(\"1:\"&LEN("+col[col_index+1].column_letter+"3))),1),\"0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ_-\"))),"
                                #only one _
                                man_formula+="ISERROR(FIND(\"_\","+col[col_index+1].column_letter+"3,FIND(\"_\","+col[col_index+1].column_letter+"3)+1)),"
                                #1st character no digit
                                man_formula+="ISERROR(LEFT("+col[col_index+1].column_letter+"3,1)*1),"
                                #last characters behind _ are numbers
                                man_formula+="NOT(ISERROR(MID("+col[col_index+1].column_letter+"3,FIND(\"_\","+col[col_index+1].column_letter+"3)+1,999)*1)),"
                                #at least one _
                                man_formula+="NOT(ISERROR(FIND(\"_\","+col[col_index+1].column_letter+"3))),"
                                # no ? (not working with white list)
                                man_formula+="ISERROR(FIND(\"?\","+col[col_index+1].column_letter+"3))"
                                man_formula+=")"
                                error_message="bug\n1.) Do not start with digit\n2.) Only use alpha-numeric charcters\n(a-Z, 0-9) and \'-\'\n3.) Replicates end with \'_1\' \'_2\'..."
                            else:
                                man_formula="=AND("
                                #only a-z A-Z 0-9 -
                                man_formula+="ISNUMBER(SUMPRODUCT(SEARCH(MID("+col[col_index+1].column_letter+"3,ROW(INDIRECT(\"1:\"&LEN("+col[col_index+1].column_letter+"3))),1),\"0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ-\"))),"
                                #1st character no digit
                                man_formula+="ISERROR(LEFT("+col[col_index+1].column_letter+"3,1)*1),"
                                # no ? (not working with white list)
                                man_formula+="ISERROR(FIND(\"?\","+col[col_index+1].column_letter+"3))"
                                man_formula+=")"
                                #man_formula="=AND(ISERROR(FIND(\"_\","+col[col_index+1].column_letter+"3)),ISERROR(FIND(\".\","+col[col_index+1].column_letter+"3)),ISERROR(LEFT("+col[col_index+1].column_letter+"3,1)*1))"
                                #man_formula=""
                                error_message="bug\n1.) Do not start with digit\n2.) Only use alpha-numeric charcters\n(a-Z, 0-9) and \'-\'"
                            #print(man_formula)
                            validation_formula=DataValidation(type="custom", formula1=man_formula, 
                                                errorStyle="stop", errorTitle="Invalid value entered", showErrorMessage=True, error=error_message)
                            ws.add_data_validation(validation_formula)
                            for row in range(3, row_num):
                                validation_formula.add(ws.cell(row=row, column=col_index+1))
                                ws.cell(row=row, column=col_index+1).protection = Protection(locked=False)
                        #auto size for sequencing name
                        if col_index+1 == user_columns-4:
                            max_length = len(cell.value)
                except:
                    pass
            if col_index+1 < user_columns:
                adjusted_width = (max_length + 3)
                ws.column_dimensions[column].width = adjusted_width
            else:
                ws.column_dimensions[column].width = 18
        ws.protection.password = excel_lock_pw
        ws.protection.sheet = True
    ws2 = wb.create_sheet(title='Contrasts')
    ws2.cell(row=1, column=1).value = "Please provide pairwise comparisons (contrasts). Keep in mind that direction of log2FC is ConditionA / ConditionB (please remove examples)"
    ws2.cell(row=2, column=1).value = "ConditionA"
    ws2.cell(row=2, column=2).value = "ConditionB"
    ws2.cell(row=3, column=1).value = "WT"
    ws2.cell(row=3, column=2).value = "Mut"

    #conditional formatting
    from openpyxl.formatting import Rule
    from openpyxl.styles.differential import DifferentialStyle
    dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='EE1111'))
    rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)
   
    for col in range(2):
        ws2.cell(row=2, column=col+1).border = Border(bottom=Side(style='medium'))
        ws2.cell(row=2, column=col+1).font = bold_font
        ws2.cell(row=2, column=col+1).alignment = align_center
        ws2.column_dimensions[ws2.cell(row=1,column=col+1).column_letter].width = 18
    cond_col=ws2.cell(row=1,column=user_columns-2).column_letter
    man_formula="="+ws.title+"!$"+str(cond_col)+"$3:$"+str(cond_col)+"$"+str(row_num-1)
    validation_list=DataValidation(type="list", formula1=man_formula, errorStyle="stop", errorTitle="Please choose valid condition from List", showErrorMessage=True)
    ws2.add_data_validation(validation_list)
    for row in range(20):
        validation_list.add(ws2.cell(row=row+3, column=1))
        validation_list.add(ws2.cell(row=row+3, column=2))
        srange="A"+str(row+3)+":B"+str(row+3)
        ws2.conditional_formatting.add(srange,rule) 
    wb.remove(wb["Sheet"]) # remove the default sheet
    wb.save(path)
    return path


def format_exp_settings(exp_data):
    exp_set_list = []
    for setting in exp_data:
        this_setting = {"organism":"", "conditions":""}
        this_setting["organism"] = format_generator(find_keys(setting, "organism_name"))[0]
        this_setting["conditions"] = format_conditions(sum(format_generator(find_keys(setting, "conditions")), []), ["sample_name"])
        exp_set_list.append(this_setting)
    return exp_set_list
    
def rename_exp_list(exp_setting_list):
    new_exp_setting_list = []
    for exp_setting in exp_setting_list:
        new_exp_setting = {}
        for condition in exp_setting["conditions"]:
            for sample_name in exp_setting["conditions"][condition]["samples"]:
                factor_dict = split_cond(condition)
                for single_sample, sample_filename in list(exp_setting["conditions"][condition]["samples"][sample_name]):
                    project_id, unique_sample_index, value_list, sample_index = split_sample_filename(sample_filename)
                    sample_label = value_list + "_" + sample_index
                    condition_label = value_list
                    new_exp_setting[single_sample] = {"condition":"", "organism_name":"", "sample_label":"", "condition_label":""}
                    new_exp_setting[single_sample]["condition"] = condition
                    new_exp_setting[single_sample]["organism_name"] = exp_setting["organism"]
                    new_exp_setting[single_sample]["sample_label"] = sample_label
                    new_exp_setting[single_sample]["condition_label"] = condition_label
                    new_exp_setting[single_sample]["sample_filename"] = sample_filename
                    new_exp_setting[single_sample]["factors"] = factor_dict

                   
        new_exp_setting_list.append(new_exp_setting)
    return new_exp_setting_list

def split_sample_filename(sample_filename):
    project_id, unique_sample_index, value_list, sample_index  = sample_filename.split("__")   
    return project_id, unique_sample_index, value_list, sample_index

def split_cond(condition):

    key_count_dict = {}
    condition_dict = {}
    conditions = []
    count = 0
    cond = '"'
    for i in range(len(condition)):
        if condition[i] == '\"':
            count += 1
            cond += condition[i]
        elif condition[i] == '-':
            if count % 2 == 0:
                conditions.append(cond)
                cond = '"'
        elif condition[i] == ':':
            cond += f'"{condition[i]}'
        elif condition[i] == '{':
            cond += f'{condition[i]}"'
        elif condition[i] == '|':
            cond += f',"'
        else:
            cond += condition[i]
    conditions.append(cond)

    for j in range(len(conditions)):
        d = json.loads(f'{"{"}{conditions[j]}{"}"}')
        key = list(d.keys())[0]
        conditions[j] = (key, d[key])
        if key not in key_count_dict.keys():
            key_count_dict[key] = 0
            dict_key = key
        else:
            key_count_dict[key] += 1
            dict_key = key + "_" + str(key_count_dict[key])
        condition_dict[dict_key] = d[key]
    return condition_dict